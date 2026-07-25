from __future__ import annotations

import csv
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import (
    ParsedDocument,
    ParsedLine,
    ParseError,
    parse_date,
    parse_decimal_field,
    parse_integer_field,
)
from ..services.normalizer import normalize_text

ALIASES = {
    "line_no": ["RIGA", "LINEA", "LINE", "N", "NUMERO LINEA"],
    "sku": ["SKU", "CODICE", "CODICE ARTICOLO", "ARTICOLO", "ITEM CODE", "PRODUCT CODE"],
    "description": ["DESCRIZIONE", "DESCRIPTION", "ARTICOLO DESCRIZIONE", "ITEM"],
    "color": ["COLORE", "COLOR", "COL"],
    "size": ["TAGLIA", "SIZE", "TG"],
    "lot": ["LOTTO", "LOT", "BATCH"],
    "unit_of_measure": ["UM", "U.M.", "UNITA MISURA", "UNIT OF MEASURE", "UOM"],
    "price_base_quantity": ["QTA BASE PREZZO", "PRICE BASE QUANTITY", "BASE QUANTITY"],
    "quantity": ["QTA", "QUANTITA", "QUANTITY", "QTY"],
    "unit_price": ["PREZZO", "PREZZO UNITARIO", "UNIT PRICE", "PRICE"],
    "discount_rate": ["SCONTO", "SCONTO PERCENTUALE", "DISCOUNT", "DISCOUNT RATE"],
    "tax_rate": ["IVA", "ALIQUOTA IVA", "TAX", "VAT"],
    "line_total": ["TOTALE", "TOTALE RIGA", "LINE TOTAL", "TOTAL"],
}


def _map_headers(headers: list[Any]) -> dict[str, int]:
    normalized = [normalize_text(str(h or "")) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            norm_alias = normalize_text(alias)
            if norm_alias in normalized:
                mapping[field] = normalized.index(norm_alias)
                break
    return mapping


def _rows(path: Path) -> list[list[Any]]:
    if path.suffix.lower() == ".csv":
        raw = path.read_text(encoding="utf-8-sig", errors="replace")
        sample = raw[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        return list(csv.reader(raw.splitlines(), dialect))
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def parse_tabular(path: Path, overrides: dict) -> ParsedDocument:
    rows = _rows(path)
    if not rows:
        raise ParseError("Il file tabellare è vuoto")
    mapping = _map_headers(rows[0])
    if "quantity" not in mapping or not ({"sku", "description"} & mapping.keys()):
        raise ParseError("Colonne minime mancanti: serve SKU o descrizione e quantità")

    doc = ParsedDocument(
        document_type=overrides.get("document_type"),
        number=overrides.get("number"),
        document_date=parse_date(overrides.get("document_date")),
        supplier_name=overrides.get("supplier_name"),
        confidence=0.92,
        metadata={"header_mapping": mapping},
    )
    if not doc.document_type:
        raise ParseError("Per CSV/XLSX è necessario indicare il tipo documento")

    for idx, row in enumerate(rows[1:], start=2):
        if not any(v not in (None, "") for v in row):
            continue

        def val(field: str):
            pos = mapping.get(field)
            return row[pos] if pos is not None and pos < len(row) else None

        line_number = parse_integer_field(val("line_no"), field="line_no", line=idx, default=idx - 1, minimum=1)
        provenance: dict[str, str] = {}
        qty = parse_decimal_field(
            val("quantity"),
            field="quantity",
            line=idx,
            required=True,
            provenance=provenance,
            max_decimal_places=4,
        )
        unit_price = parse_decimal_field(
            val("unit_price"),
            field="unit_price",
            line=idx,
            provenance=provenance,
            max_decimal_places=6,
        )
        discount = parse_decimal_field(
            val("discount_rate"),
            field="discount_rate",
            line=idx,
            provenance=provenance,
            max_decimal_places=6,
            minimum=0,
            maximum=100,
        )
        base_qty = parse_decimal_field(
            val("price_base_quantity"),
            field="price_base_quantity",
            line=idx,
            default=1,
            provenance=provenance,
            missing_provenance="defaulted",
            max_decimal_places=4,
            exclusive_minimum=0,
        )
        tax = parse_decimal_field(
            val("tax_rate"),
            field="tax_rate",
            line=idx,
            provenance=provenance,
            max_decimal_places=6,
            minimum=0,
            maximum=100,
        )
        expected = qty * unit_price / base_qty * (Decimal("1") - discount / Decimal("100"))
        if val("line_total") in (None, ""):
            total = expected
            provenance["line_total"] = "derived" if provenance.get("unit_price") == "source" else "missing"
        else:
            total = parse_decimal_field(
                val("line_total"),
                field="line_total",
                line=idx,
                required=True,
                provenance=provenance,
                max_decimal_places=2,
            )
        doc.lines.append(
            ParsedLine(
                line_no=line_number,
                sku=str(val("sku")).strip() if val("sku") not in (None, "") else None,
                description=str(val("description")).strip() if val("description") not in (None, "") else None,
                color=str(val("color")).strip() if val("color") not in (None, "") else None,
                size=str(val("size")).strip() if val("size") not in (None, "") else None,
                lot=str(val("lot")).strip() if val("lot") not in (None, "") else None,
                unit_of_measure=(
                    str(val("unit_of_measure")).strip() if val("unit_of_measure") not in (None, "") else None
                ),
                quantity=qty,
                unit_price=unit_price,
                price_base_quantity=base_qty,
                discount_rate=discount,
                tax_rate=tax,
                line_total=total,
                confidence=0.94,
                raw={"row": idx, "numeric_provenance": provenance},
            )
        )
    if not doc.lines:
        raise ParseError("Nessuna riga dati valida trovata")
    return doc
